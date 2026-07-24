from __future__ import annotations

from io import BytesIO
from typing import Any

from openpyxl import Workbook, load_workbook


# Возможные названия колонок в файле Excel (регистр не важен)
HEADER_ALIASES: dict[str, list[str]] = {
    "full_name": [
        "фио",
        "ф.и.о.",
        "ф.и.о",
        "имя",
        "студент",
        "full_name",
        "fullname",
        "fio",
    ]
}


def _norm(value: Any) -> str:
    """Приводит значение ячейки к строке без лишних пробелов."""
    if value is None:
        return ""
    return str(value).strip()


def _match_header(cell: Any) -> str | None:
    """Определяет, к какому полю относится заголовок колонки."""
    text = _norm(cell).lower().replace("ё", "е")
    if not text:
        return None
    for field, aliases in HEADER_ALIASES.items():
        for alias in aliases:
            if text == alias or alias in text:
                return field
    return None


def parse_students_xlsx(data: bytes) -> tuple[list[dict[str, Any]], str | None]:
    """
    Разбирает Excel-файл и возвращает список студентов.

    Ожидаемые колонки: ФИО.

    Возвращает: (список_словарей, текст_ошибки_или_None).
    """
    try:
        # Открываем книгу Excel из байтов загруженного файла
        wb = load_workbook(BytesIO(data), data_only=True)
    except Exception as exc: 
        return [], f"Не удалось прочитать Excel: {exc}"

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], "Файл пустой"

    header_map: dict[str, int] = {}
    start_idx = 0
    first = rows[0]

    # Пытаемся распознать строку заголовков
    for col_idx, cell in enumerate(first):
        field = _match_header(cell)
        if field and field not in header_map:
            header_map[field] = col_idx

    if "full_name" in header_map:
        start_idx = 1  # данные начинаются со второй строки
    else:
        # Без заголовков: A = ФИО
        header_map = {"full_name": 0}
        start_idx = 0

    students: list[dict[str, Any]] = []
    for row in rows[start_idx:]:
        # Пропускаем пустые строки
        if row is None or all(cell is None or str(cell).strip() == "" for cell in row):
            continue

        full_name = _norm(row[header_map["full_name"]]) if header_map["full_name"] < len(row) else ""
  

        students.append(
            {
                "full_name": full_name
            }
        )

    if not students:
        return [], "В файле не найдено ни одной строки со студентами"

    return students, None


def build_template_xlsx() -> bytes:
    """Создаёт шаблон Excel для импорта группы (пример заполнения)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Студенты"
    ws.append(["ФИО"])
    ws.append(["Иванов Иван Иванович"])
    ws.append(["Петрова Анна Сергеевна"])
    ws.append(["Сидоров Пётр Алексеевич"])
    for col in ("A"):
        ws.column_dimensions[col].width = 28 if col != "C" else 10
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
